"""
Export mart data to a formatted Excel report.
Output: output/ph_fx_report.xlsx

Usage:
    python scripts/export_excel.py
"""

from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DSN     = os.environ.get("PH_FX_POSTGRES_DSN", "postgresql://fx:fx@localhost:5432/ph_fx")
OUT_DIR = Path(__file__).parent.parent / "output"
OUT_DIR.mkdir(exist_ok=True)
OUT_PATH = OUT_DIR / "ph_fx_report.xlsx"

HEADER_FILL  = PatternFill("solid", fgColor="0D1B2A")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")


def fetch(query: str) -> pd.DataFrame:
    with psycopg2.connect(DSN) as conn:
        return pd.read_sql(query, conn)


def style_sheet(ws, df: pd.DataFrame) -> None:
    """Apply header styling and column auto-width."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(ws.max_row + 1, 100))
        )
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 40))


def main() -> None:
    print("Fetching mart data...")

    monthly = fetch("""
        SELECT
            DATE_TRUNC('month', rate_date)::DATE AS month,
            ROUND(AVG(rate)::NUMERIC, 4)         AS avg_rate,
            ROUND(MIN(rate)::NUMERIC, 4)         AS min_rate,
            ROUND(MAX(rate)::NUMERIC, 4)         AS max_rate,
            ROUND(AVG(vol_30d)::NUMERIC, 6)      AS avg_volatility
        FROM marts.fx_dashboard
        JOIN marts.fx_volatility USING (rate_date)
        GROUP BY 1
        ORDER BY 1
    """)

    dashboard = fetch("""
        SELECT rate_date, rate, avg_7d, avg_30d, ytd_low, ytd_high,
               daily_change_pct, change_30d_pct
        FROM marts.fx_dashboard
        ORDER BY rate_date DESC
        LIMIT 365
    """)

    real_rate = fetch("""
        SELECT month, nominal_rate, real_rate, inflation_gap, inflation_pct
        FROM marts.real_exchange_rate
        ORDER BY month DESC
        LIMIT 60
    """)

    cross = fetch("""
        SELECT base_currency, php_rate, rate_date
        FROM raw.cross_rates
        WHERE rate_date = (SELECT MAX(rate_date) FROM raw.cross_rates)
        ORDER BY base_currency
    """)

    print(f"Writing to {OUT_PATH}...")
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        monthly.to_excel(writer,   sheet_name="Monthly Summary", index=False)
        dashboard.to_excel(writer, sheet_name="Daily Rates (Last 365d)", index=False)
        real_rate.to_excel(writer, sheet_name="Real Exchange Rate", index=False)
        cross.to_excel(writer,     sheet_name="Cross Rates", index=False)

    wb = load_workbook(OUT_PATH)
    for sheet_name, df in [
        ("Monthly Summary", monthly),
        ("Daily Rates (Last 365d)", dashboard),
        ("Real Exchange Rate", real_rate),
        ("Cross Rates", cross),
    ]:
        style_sheet(wb[sheet_name], df)
    wb.save(OUT_PATH)

    size_kb = OUT_PATH.stat().st_size // 1024
    print(f"Done → {OUT_PATH}  ({size_kb:,} KB)")


if __name__ == "__main__":
    main()
